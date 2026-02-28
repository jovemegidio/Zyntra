#!/usr/bin/env python3
"""
Analyze and fix Lista de Estoque - Aluforce Cabos.xlsx
- Fix missing values in 'Quantidade em estoque (bobinas)' column
- Ensure bobina/rolo classification is correct
- Validate all data consistency
- Generate updated Excel file
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import warnings
warnings.filterwarnings('ignore')

FILE_PATH = r'G:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\Arvore de Produto com Custo\Lista de Estoque - Aluforce Cabos.xlsx'

def analyze():
    """Analyze the Excel file and report issues"""
    df = pd.read_excel(FILE_PATH, sheet_name='Lista de estoque', header=3)
    qtde_col = df.columns[4]  # 'Quantidade em\nestoque (bobinas'
    
    print("=" * 70)
    print("ANÁLISE DO ARQUIVO: Lista de Estoque - Aluforce Cabos.xlsx")
    print("=" * 70)
    print(f"\nTotal de linhas: {len(df)}")
    print(f"Colunas: {list(df.columns)}")
    print()
    
    # Issue 1: NaN in Quantidade em estoque
    nan_qtde = df[df[qtde_col].isna()]
    print(f"PROBLEMA 1: {len(nan_qtde)} linhas SEM valor em '{qtde_col}'")
    for idx, row in nan_qtde.iterrows():
        print(f"  Linha Excel {idx+5}: COD={row['COD']}, Nome={row['Nome']}, QTDE={row['QTDE']}m")
    print()
    
    # Issue 2: NaN in Bobinas column (no dimension or ROLO indicator)
    nan_bob = df[df['Bobinas'].isna()]
    print(f"PROBLEMA 2: {len(nan_bob)} linhas sem info na coluna 'Bobinas'")
    print("  (São bobinas sem dimensão registrada - deveria ter 'BOBINA' ou dimensão)")
    print()
    
    # Classification
    def classify(b):
        if pd.isna(b): return 'BOBINA'
        if str(b).strip().upper() == 'ROLO': return 'ROLO'
        return 'BOBINA'
    
    df['TIPO'] = df['Bobinas'].apply(classify)
    
    bobinas = df[df['TIPO'] == 'BOBINA']
    rolos = df[df['TIPO'] == 'ROLO']
    
    print(f"CLASSIFICAÇÃO:")
    print(f"  Bobinas: {len(bobinas)} (com dimensão: {len(bobinas[bobinas['Bobinas'].notna()])})")
    print(f"  Rolos: {len(rolos)}")
    print(f"  Total: {len(df)}")
    print()
    
    # Summary by code
    print("RESUMO POR CÓDIGO:")
    for cod in sorted(df['COD'].unique()):
        sub = df[df['COD'] == cod]
        bobs = len(sub[sub['TIPO'] == 'BOBINA'])
        rolos_count = len(sub[sub['TIPO'] == 'ROLO'])
        total_m = sub['QTDE'].sum()
        items = []
        if bobs > 0:
            items.append(f"{bobs} bobina(s)")
        if rolos_count > 0:
            items.append(f"{rolos_count} rolo(s)")
        print(f"  {cod:10s}: {' + '.join(items):25s} = {bobs+rolos_count} unid. ({total_m:,.0f}m)")
    
    return df

def fix_and_save():
    """Fix issues and save updated Excel file"""
    print("\n" + "=" * 70)
    print("ATUALIZANDO ARQUIVO EXCEL")
    print("=" * 70)
    
    # Load workbook preserving formatting
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Lista de estoque']
    
    # Read data with pandas for analysis
    df = pd.read_excel(FILE_PATH, sheet_name='Lista de estoque', header=3)
    qtde_col = df.columns[4]
    
    # Header row is row 4 (1-indexed in openpyxl)
    HEADER_ROW = 4
    DATA_START_ROW = 5
    
    # Column indices (1-indexed for openpyxl)
    COL_COD = 1       # A
    COL_NOME = 2      # B
    COL_QTDE = 3      # C
    COL_BOBINAS = 4   # D - Dimensão/Tipo
    COL_QTDE_EST = 5  # E - Quantidade em estoque (bobinas)
    COL_COR = 6       # F
    COL_LOCAL = 7      # G
    COL_OBS = 8        # H
    
    fixes_applied = 0
    
    # Fix 1: Fill NaN in "Quantidade em estoque (bobinas)" with 1
    for idx, row in df.iterrows():
        excel_row = DATA_START_ROW + idx
        cell = ws.cell(row=excel_row, column=COL_QTDE_EST)
        
        if cell.value is None or (isinstance(cell.value, float) and pd.isna(cell.value)):
            cell.value = 1
            # Copy formatting from adjacent row
            ref_cell = ws.cell(row=DATA_START_ROW, column=COL_QTDE_EST)
            cell.font = copy(ref_cell.font) if ref_cell.font else Font()
            cell.alignment = copy(ref_cell.alignment) if ref_cell.alignment else Alignment()
            cell.number_format = ref_cell.number_format
            if ref_cell.fill:
                cell.fill = copy(ref_cell.fill)
            if ref_cell.border:
                cell.border = copy(ref_cell.border)
            fixes_applied += 1
            print(f"  FIX: Linha {excel_row} ({row['COD']}) - Quantidade em estoque definida para 1")
    
    # Fix 2: For rows where Bobinas is NaN (no dimension), set to 'BOBINA' to differentiate from ROLO
    for idx, row in df.iterrows():
        excel_row = DATA_START_ROW + idx
        cell = ws.cell(row=excel_row, column=COL_BOBINAS)
        
        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
            # It's a bobina without dimension info
            cell.value = 'BOBINA'
            ref_cell = ws.cell(row=DATA_START_ROW, column=COL_BOBINAS)
            cell.font = copy(ref_cell.font) if ref_cell.font else Font()
            cell.alignment = copy(ref_cell.alignment) if ref_cell.alignment else Alignment(horizontal='center')
            cell.number_format = ref_cell.number_format
            if ref_cell.fill:
                cell.fill = copy(ref_cell.fill)
            if ref_cell.border:
                cell.border = copy(ref_cell.border)
            fixes_applied += 1
    
    print(f"\n  Total de correções: {fixes_applied}")
    
    # Save
    wb.save(FILE_PATH)
    print(f"\n  ✅ Arquivo salvo: {FILE_PATH}")
    
    # Verify
    print("\n" + "=" * 70)
    print("VERIFICAÇÃO PÓS-ATUALIZAÇÃO")
    print("=" * 70)
    
    df2 = pd.read_excel(FILE_PATH, sheet_name='Lista de estoque', header=3)
    qtde_col2 = df2.columns[4]
    
    nan_qtde = df2[df2[qtde_col2].isna()]
    print(f"\n  Linhas sem Quantidade em estoque: {len(nan_qtde)}")
    
    nan_bob = df2[df2['Bobinas'].isna()]
    print(f"  Linhas sem classificação Bobinas: {len(nan_bob)}")
    
    # Count types
    def classify(b):
        if pd.isna(b): return 'SEM TIPO'
        b_str = str(b).strip().upper()
        if b_str == 'ROLO': return 'ROLO'
        if b_str == 'BOBINA': return 'BOBINA'
        if 'X' in b_str or ',' in b_str: return 'BOBINA (com dimensão)'
        return f'OUTRO: {b_str}'
    
    df2['TIPO'] = df2['Bobinas'].apply(classify)
    print(f"\n  Distribuição por tipo:")
    for tipo, count in df2['TIPO'].value_counts().items():
        print(f"    {tipo}: {count}")
    
    total_bobinas = len(df2[df2['TIPO'].str.contains('BOBINA', na=False)])
    total_rolos = len(df2[df2['TIPO'] == 'ROLO'])
    print(f"\n  TOTAL BOBINAS: {total_bobinas}")
    print(f"  TOTAL ROLOS: {total_rolos}")
    print(f"  TOTAL GERAL: {total_bobinas + total_rolos}")
    print(f"  TOTAL METROS: {df2['QTDE'].sum():,.0f}m")

if __name__ == '__main__':
    analyze()
    fix_and_save()
