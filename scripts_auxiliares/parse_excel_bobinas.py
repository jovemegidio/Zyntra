#!/usr/bin/env python3
"""Parse Excel file and generate SQL import for bobinas_estoque table."""
import openpyxl
import json

EXCEL_PATH = r'g:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\Arvore de Produto com Custo\Lista de Estoque - Aluforce Cabos.xlsx'
OUTPUT_JSON = r'g:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\scripts_auxiliares\import_bobinas.json'
OUTPUT_SQL = r'g:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\scripts_auxiliares\import_bobinas.sql'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Lista de estoque']

all_rows = []
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    vals = (list(row) + [None]*8)[:8]
    cod, nome, qtde, bobina_dim, qtde_bob, veia_cor, local_str, obs = vals
    if cod and qtde and str(cod).strip().upper() != 'COD' and str(qtde).strip().upper() != 'QTDE':
        try:
            qtde_val = float(qtde)
        except (ValueError, TypeError):
            continue
        all_rows.append({
            'cod': str(cod).strip(),
            'nome': str(nome).strip() if nome else '',
            'qtde': qtde_val,
            'bobina_dim': str(bobina_dim).strip() if bobina_dim else None,
            'qtde_bob': int(qtde_bob) if qtde_bob else 1,
            'veia_cor': str(veia_cor).strip() if veia_cor else None,
            'local': str(local_str).strip() if local_str else None,
            'obs': str(obs).strip() if obs else None
        })

print(f"Total data rows: {len(all_rows)}")

# Group by cod
by_cod = {}
for r in all_rows:
    if r['cod'] not in by_cod:
        by_cod[r['cod']] = {'nome': r['nome'], 'bobinas': []}
    by_cod[r['cod']]['bobinas'].append(r)

print(f"Unique product codes: {len(by_cod)}")

# Save JSON
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(all_rows, f, ensure_ascii=False, indent=2)

# Generate SQL
def esc(val):
    if val is None:
        return 'NULL'
    s = str(val).replace("'", "\\'").replace("\\", "\\\\")
    return f"'{s}'"

sql_lines = []
sql_lines.append("-- Import bobinas from Excel - Stock zeroed (estoque_atual = 0)")
sql_lines.append("-- Generated automatically from Lista de Estoque - Aluforce Cabos.xlsx")
sql_lines.append("")
sql_lines.append("-- Step 1: Ensure all product codes exist in produtos table")
sql_lines.append("")

for cod, data in by_cod.items():
    nome = data['nome']
    # Get first bobina's color for variacao
    first_cor = data['bobinas'][0]['veia_cor'] or ''
    total_qtde = sum(b['qtde'] for b in data['bobinas'])
    total_bobinas = len(data['bobinas'])
    
    nome_esc = esc(nome)
    cod_esc = esc(cod)
    cor_esc = esc(first_cor) if first_cor else 'NULL'
    
    sql_lines.append(f"INSERT INTO produtos (codigo, nome, descricao, categoria, estoque_atual, quantidade_estoque, estoque_minimo, unidade_medida, cor, status, ativo)")
    sql_lines.append(f"VALUES ({cod_esc}, {nome_esc}, {nome_esc}, 'CABOS', 0, 0, 5, 'M', {cor_esc}, 'ativo', 1)")
    sql_lines.append(f"ON DUPLICATE KEY UPDATE nome = VALUES(nome), descricao = VALUES(descricao);")
    sql_lines.append("")

sql_lines.append("")
sql_lines.append("-- Step 2: Insert bobinas (stock is 0 for display, bobinas have the real data)")
sql_lines.append("")

bobina_num_tracker = {}
for r in all_rows:
    cod = r['cod']
    if cod not in bobina_num_tracker:
        bobina_num_tracker[cod] = 0
    bobina_num_tracker[cod] += 1
    num = bobina_num_tracker[cod]
    
    sql_lines.append(
        f"INSERT INTO bobinas_estoque (produto_id, codigo_produto, quantidade, dimensao_bobina, veia_cor, local_armazenamento, observacao, status, numero_bobina)"
        f" SELECT id, {esc(cod)}, {r['qtde']}, {esc(r['bobina_dim'])}, {esc(r['veia_cor'])}, {esc(r['local'])}, {esc(r['obs'])}, 'disponivel', {num}"
        f" FROM produtos WHERE codigo = {esc(cod)} LIMIT 1;"
    )

sql_lines.append("")
sql_lines.append("-- Step 3: Verify import")
sql_lines.append("SELECT p.codigo, p.nome, COUNT(b.id) as total_bobinas, SUM(b.quantidade) as total_metros")
sql_lines.append("FROM produtos p JOIN bobinas_estoque b ON b.produto_id = p.id")
sql_lines.append("GROUP BY p.codigo, p.nome ORDER BY p.codigo;")

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))

print(f"SQL saved to: {OUTPUT_SQL}")
print(f"JSON saved to: {OUTPUT_JSON}")

# Print summary
for cod, data in sorted(by_cod.items()):
    total = sum(b['qtde'] for b in data['bobinas'])
    bcount = len(data['bobinas'])
    print(f"  {cod}: {data['nome'][:50]} - {bcount} bobina(s), total={total}m")
