import openpyxl
import json

wb = openpyxl.load_workbook(
    r'G:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\Arvore de Produto com Custo\Lista de Estoque - Aluforce Cabos.xlsx',
    data_only=True
)
ws = wb['Lista de estoque']

print(f"Total linhas: {ws.max_row}")
print()

# Headers row 4
headers = []
for col in range(1, 9):
    h = ws.cell(4, col).value
    headers.append(str(h or '').strip())
print("Headers:", headers)
print()

# Collect all data rows (data starts row 5)
all_rows = []
for row_num in range(5, ws.max_row + 1):
    a = ws.cell(row_num, 1).value  # COD
    b = ws.cell(row_num, 2).value  # Nome
    c = ws.cell(row_num, 3).value  # QTDE (metros)
    d = ws.cell(row_num, 4).value  # Bobinas (dimensao)
    e = ws.cell(row_num, 5).value  # Qtd em estoque (bobinas)
    f = ws.cell(row_num, 6).value  # VEIA / COR
    g = ws.cell(row_num, 7).value  # LOCAL
    h = ws.cell(row_num, 8).value  # Observacao

    if a is None and b is None and c is None:
        continue
    
    try:
        qtde_val = float(c) if c is not None else 0
    except (ValueError, TypeError):
        qtde_val = 0
    
    try:
        bob_val = int(e) if e is not None else 0
    except (ValueError, TypeError):
        bob_val = 0

    all_rows.append({
        'row': row_num,
        'cod': str(a or '').strip(),
        'nome': str(b or '').strip(),
        'qtde': qtde_val,
        'dimensao': str(d or '').strip(),
        'qtd_bobinas': bob_val,
        'cor': str(f or '').strip(),
        'local': str(g or '').strip(),
        'obs': str(h or '').strip()
    })

print(f"Total registros (linhas com dados): {len(all_rows)}")
print()

# Print each row
for r in all_rows:
    cod = r['cod'][:10].ljust(10)
    nome = r['nome'][:50].ljust(50)
    qtde = str(r['qtde']).ljust(8)
    dim = r['dimensao'][:14].ljust(14)
    qb = str(r['qtd_bobinas']).ljust(5)
    cor = r['cor'][:12].ljust(12)
    loc = r['local'][:20].ljust(20)
    obs = r['obs'][:25].ljust(25)
    print(f"  {cod} {nome} {qtde} {dim} {qb} {cor} {loc} {obs}")

# Summary by product code
print()
print("=" * 80)
print("RESUMO POR CODIGO:")
print("=" * 80)

from collections import defaultdict
by_code = defaultdict(list)
for r in all_rows:
    if r['cod']:
        by_code[r['cod']].append(r)

for cod in sorted(by_code.keys()):
    rows = by_code[cod]
    total_metros = sum(r['qtde'] for r in rows)
    total_bobinas = len(rows)
    cores = set(r['cor'] for r in rows if r['cor'])
    locais = set(r['local'] for r in rows if r['local'])
    dims = set(r['dimensao'] for r in rows if r['dimensao'])
    obs_list = set(r['obs'] for r in rows if r['obs'])
    
    print(f"  {cod:<10} | {total_bobinas:>3} linhas | {total_metros:>8.0f}m | dims: {', '.join(sorted(dims)):<25} | cores: {', '.join(sorted(cores)):<15} | locais: {', '.join(sorted(locais))}")
    if obs_list:
        for o in obs_list:
            print(f"{'':>17}obs: {o}")

# Export as JSON for SQL generation
print()
print("=== JSON DATA ===")
print(json.dumps(all_rows, ensure_ascii=False, indent=None))
