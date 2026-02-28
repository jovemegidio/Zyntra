import openpyxl
import warnings
import os

warnings.filterwarnings('ignore')

base = r'G:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\templates'
files = [
    'Omie_Clientes_Fornecedores_v1_5_16.xlsx',
    'Omie_Contas_Pagar_v1_1_5.xlsx',
    'Omie_Contas_Receber_v1_0_6.xlsx',
    'Omie_Contratos_v1_4_2.xlsx',
    'Omie_Produtos_v1_9_5.xlsx',
    'Omie_Servicos_v1_0_10.xlsx',
    'Omie_Importacao_Mercadoria_v1_0_8_5.xlsx',
    'Omie_Ordens_Servico_v1_2_12.xlsx',
    'Omie_Remessa_Produto_v1_2_2.xlsx'
]

for f in files:
    path = os.path.join(base, f)
    wb = openpyxl.load_workbook(path)
    print(f'=== {f} ===')
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f'  Sheet: {ws_name} ({ws.max_row}r x {ws.max_column}c)')
        # Check images
        imgs = ws._images
        print(f'  Images: {len(imgs)}')
        for idx, img in enumerate(imgs):
            try:
                anc = type(img.anchor).__name__
                print(f'    - img[{idx}]: type={anc}, w={img.width}, h={img.height}')
                if hasattr(img.anchor, '_from'):
                    f = img.anchor._from
                    print(f'      from: col={f.col}, row={f.row}')
                if hasattr(img.anchor, 'col'):
                    print(f'      col={img.anchor.col}, row={img.anchor.row}')
            except Exception as e:
                print(f'    - img[{idx}]: error reading: {e}')
        # Read first 10 rows, first 5 cols
        for r in range(1, min(11, ws.max_row + 1)):
            vals = []
            for c in range(1, min(6, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v is not None:
                    s = str(v)[:50]
                    vals.append(f'C{c}={repr(s)}')
            if vals:
                sep = ' | '
                print(f'    Row {r}: {sep.join(vals)}')
        # Merged cells
        merges = list(ws.merged_cells.ranges)
        if merges:
            print(f'  Merged: {merges[:5]}')
        # Row heights for first rows
        for r in range(1, 10):
            h = ws.row_dimensions[r].height
            if h:
                print(f'    Row {r} height: {h}')
    print()
    wb.close()

# Check Zyntra logo
from PIL import Image
logo_path = r'G:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2\Zyntra\Zyntra - Branco.png'
try:
    img = Image.open(logo_path)
    print(f'=== ZYNTRA LOGO ===')
    print(f'  Size: {img.size}')
    print(f'  Mode: {img.mode}')
    print(f'  Format: {img.format}')
except Exception as e:
    print(f'Logo error: {e}')
