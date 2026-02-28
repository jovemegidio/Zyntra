"""
Script para recriar os templates Omie com o logo Zyntra
Mantém toda a estrutura original do Omie (formatação, validações, Config, etc.)
Apenas substitui o logo Omie pelo logo Zyntra em todas as abas
"""
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from PIL import Image as PILImage
from copy import copy
import warnings
import shutil
import os
import tempfile

warnings.filterwarnings('ignore')

# Paths
BASE = r'G:\Outros computadores\Meu laptop (2)\Sistema - ALUFORCE - V.2'
TEMPLATES_DIR = os.path.join(BASE, 'templates')
OUTPUT_DIR = os.path.join(TEMPLATES_DIR, 'zyntra')
LOGO_PATH = os.path.join(BASE, 'Zyntra', 'Zyntra - Branco.png')

# Pre-resize the Zyntra logo to common Omie dimensions (279x148)
# This ensures openpyxl uses correct native pixel dimensions
RESIZED_LOGOS = {}  # cache: (w,h) -> temp file path

def get_resized_logo(target_w, target_h):
    """Get a pre-resized Zyntra logo PNG matching the target dimensions."""
    key = (int(target_w), int(target_h))
    if key in RESIZED_LOGOS:
        return RESIZED_LOGOS[key]
    
    img = PILImage.open(LOGO_PATH)
    img = img.resize(key, PILImage.LANCZOS)
    
    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False, 
                                       dir=OUTPUT_DIR, prefix=f'_logo_{key[0]}x{key[1]}_')
    img.save(tmp.name, 'PNG')
    tmp.close()
    RESIZED_LOGOS[key] = tmp.name
    return tmp.name

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Validate prerequisites
if not os.path.isfile(LOGO_PATH):
    print(f'ERRO: Logo Zyntra não encontrado: {LOGO_PATH}')
    exit(1)

# All Omie template files
OMIE_FILES = [
    'Omie_Clientes_Fornecedores_v1_5_16.xlsx',
    'Omie_Contas_Pagar_v1_1_5.xlsx',
    'Omie_Contas_Receber_v1_0_6.xlsx',
    'Omie_Contratos_v1_4_2.xlsx',
    'Omie_Produtos_v1_9_5.xlsx',
    'Omie_Servicos_v1_0_10.xlsx',
    'Omie_Importacao_Mercadoria_v1_0_8_5.xlsx',
    'Omie_Ordens_Servico_v1_2_12.xlsx',
    'Omie_Remessa_Produto_v1_2_2.xlsx',
]

# Validate all source templates exist
missing = [f for f in OMIE_FILES if not os.path.isfile(os.path.join(TEMPLATES_DIR, f))]
if missing:
    print(f'AVISO: {len(missing)} templates Omie não encontrados:')
    for m in missing:
        print(f'  - {m}')
    OMIE_FILES = [f for f in OMIE_FILES if f not in missing]

def replace_logo_in_template(omie_filename):
    """
    Opens an Omie template, removes all images from data sheets,
    and inserts the Zyntra logo in the same position.
    Keeps Config sheet untouched.
    """
    src_path = os.path.join(TEMPLATES_DIR, omie_filename)
    # Output with same name (in zyntra subfolder)
    out_filename = omie_filename.replace('Omie_', 'Zyntra_')
    out_path = os.path.join(OUTPUT_DIR, out_filename)
    
    # First copy the file as-is
    shutil.copy2(src_path, out_path)
    
    # Now open and modify
    wb = openpyxl.load_workbook(out_path)
    
    sheets_modified = 0
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # Skip Config sheets - no images to replace
        if ws_name == 'Config':
            continue
        
        old_images = ws._images[:]
        if not old_images:
            continue
        
        # Collect info about existing images (positions)
        image_info = []
        for img in old_images:
            try:
                if hasattr(img.anchor, '_from'):
                    from_marker = img.anchor._from
                    col = from_marker.col
                    row = from_marker.row
                else:
                    col = 0
                    row = 0
                image_info.append({
                    'col': col,
                    'row': row,
                    'width': img.width,
                    'height': img.height
                })
            except Exception:
                image_info.append({
                    'col': 0,
                    'row': 0,
                    'width': 279,
                    'height': 148
                })
        
        # Clear all images
        ws._images = []
        
        # Add Zyntra logo for each position where there was an Omie logo
        # Only replace the main logo (largest one, usually ~279x148, at row 0)
        logo_added = False
        for info in image_info:
            if info['row'] == 0 and not logo_added:
                # This is the main Omie logo position - replace with Zyntra
                # Use pre-resized logo to match exact Omie dimensions
                target_w = max(int(info['width']), 100)   # pixels from original
                target_h = max(int(info['height']), 50)   # pixels from original
                resized_path = get_resized_logo(target_w, target_h)
                zyntra_img = XlImage(resized_path)
                
                # Position at same cell
                col_letter = openpyxl.utils.get_column_letter(info['col'] + 1)
                cell_ref = f'{col_letter}1'
                ws.add_image(zyntra_img, cell_ref)
                logo_added = True
                sheets_modified += 1
            # Skip the smaller secondary images (154x44) that appear in some sheets
            # These seem to be duplicate/artifacts, not needed
    
    wb.save(out_path)
    wb.close()
    
    print(f'  OK: {out_filename} ({sheets_modified} abas com logo)')
    return out_filename

# Process all templates
print('Recriando templates com logo Zyntra...\n')

created_files = []
for omie_file in OMIE_FILES:
    try:
        out_name = replace_logo_in_template(omie_file)
        created_files.append(out_name)
    except Exception as e:
        print(f'  ERRO: {omie_file} -> {e}')

print(f'\n{len(created_files)} templates criados em: {OUTPUT_DIR}')

# Clean up temp resized logos
for tmp_path in RESIZED_LOGOS.values():
    try:
        os.unlink(tmp_path)
    except OSError:
        pass

print(f'\nArquivos finais em {OUTPUT_DIR}:')
for f in sorted(os.listdir(OUTPUT_DIR)):
    if f.endswith('.xlsx'):
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        print(f'  {f} ({size/1024:.0f} KB)')
